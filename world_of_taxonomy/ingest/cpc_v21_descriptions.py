"""Parser for the UN CPC v2.1 explanatory notes spreadsheet.

The structural ingester at :mod:`world_of_taxonomy.ingest.cpc_v21`
loads the 4,596-code hierarchy from the plain-text structure file
(`CPC_Ver_2_1_english_structure.txt`), which carries only code +
title. UNSD also publishes a separate XLSX
(`CPC_Ver_2.1_Exp_Notes_Updated_<date>.xlsx`) with per-code
inclusion and exclusion paragraphs. This parser surfaces those
paragraphs as a markdown-formatted description.

Sheet layout (UNSD, April 2025 release):
  Sheet "CPC2.1": Code | Title | Inclusions | Exclusions  (one row per code)
  Sheet "61_62":  *** wildcard codes for trade services divisions
                  (Code | Title | Inclusions only). Each ``***NN`` code
                  expands to every 5-digit class in divisions 61 and 62
                  whose final three digits match ``NN``. We expand the
                  wildcards to concrete codes for direct DB matching.

Returns ``{code: markdown}`` for every code that has at least one
inclusion or exclusion paragraph; codes with neither are omitted so
the backfill does not write empty strings.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl


_INCLUDES_HEADER = "**Includes:**"
_EXCLUDES_HEADER = "**Excludes:**"


def parse_cpc_v21_exp_notes_xlsx(xlsx_path: Path) -> Dict[str, str]:
    """Return ``{code: markdown_description}`` from the UNSD XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: Dict[str, str] = {}

    out.update(_parse_main_sheet(wb["CPC2.1"]))
    if "61_62" in wb.sheetnames:
        out.update(_parse_wildcard_sheet(wb["61_62"]))

    return out


def _parse_main_sheet(ws) -> Dict[str, str]:
    """Parse the primary CPC2.1 sheet (Code | Title | Includes | Excludes)."""
    out: Dict[str, str] = {}
    for code, title, includes, excludes in _data_rows(ws, expected_cols=4):
        body = _render_markdown(includes, excludes)
        if body:
            out[code] = body
    return out


def _parse_wildcard_sheet(ws) -> Dict[str, str]:
    """Parse the 61_62 sheet and expand ``***NN`` wildcards.

    Codes in this sheet are like ``***1`` (4-digit) and ``***11`` (5-digit).
    They expand to every concrete code under divisions 61 and 62 whose
    last digits match. We emit the wildcard's text against each expanded
    code; the structural ingester guarantees the codes already exist.
    """
    out: Dict[str, str] = {}
    for code, title, includes in _data_rows(ws, expected_cols=3):
        if not code.startswith("***"):
            continue
        suffix = code[3:]
        body = _render_markdown(includes, None)
        if not body:
            continue
        for division in ("61", "62"):
            expanded = f"{division}{suffix}"
            out[expanded] = body
    return out


def _data_rows(
    ws, *, expected_cols: int,
) -> Iterable[Tuple[str, ...]]:
    """Yield data rows (skipping headers and notes), padding to expected_cols."""
    seen_header = False
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = row[0]
        if first is None:
            continue
        first_str = str(first).strip()
        if not first_str:
            continue
        # Header row contains "CPC Ver. 2.1 Code"
        if "Code" in first_str and "CPC" in first_str:
            seen_header = True
            continue
        if not seen_header:
            continue
        # Data row: pad/truncate to expected_cols
        cells = list(row[:expected_cols]) + [None] * max(
            0, expected_cols - len(row),
        )
        cells = [str(c).strip() if c is not None else "" for c in cells]
        yield tuple(cells)


def _render_markdown(includes: Optional[str], excludes: Optional[str]) -> str:
    """Combine inclusion and exclusion paragraphs into a markdown body.

    The XLSX cells contain bullet-style text already (``- item`` or ``* item``);
    we keep their structure but lead with bold headers so the result is
    consistent with other description-backfill outputs.
    """
    blocks: List[str] = []
    inc = _normalize(includes)
    if inc:
        blocks.append(f"{_INCLUDES_HEADER}\n{_strip_leading_label(inc)}")
    exc = _normalize(excludes)
    if exc:
        blocks.append(f"{_EXCLUDES_HEADER}\n{_strip_leading_label(exc)}")
    return "\n\n".join(blocks)


def _normalize(s: Optional[str]) -> str:
    if s is None:
        return ""
    if not s.strip():
        return ""
    # Collapse runs of blank lines but preserve the bullet structure
    lines = [line.rstrip() for line in s.splitlines()]
    return "\n".join(line for line in lines if line.strip())


def _strip_leading_label(s: str) -> str:
    """Remove the verbose leading sentence UNSD prepends to most cells.

    The XLSX cells typically start with "This subclass includes:" or
    "This group includes:" before the bullet list. We strip those
    leading sentences so the markdown reads cleanly under the bold
    header.
    """
    lines = s.splitlines()
    if not lines:
        return s
    first = lines[0].strip().lower()
    leading_phrases = (
        "this subclass includes:",
        "this subclass does not include:",
        "this class includes:",
        "this class does not include:",
        "this group includes:",
        "this group does not include:",
        "this division includes:",
        "this division does not include:",
        "this section includes:",
        "this section does not include:",
    )
    if first in leading_phrases:
        return "\n".join(lines[1:]).strip()
    return s.strip()
